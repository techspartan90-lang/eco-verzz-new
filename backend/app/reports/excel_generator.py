import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    """
    OpenPyXL Excel Generator creating multi-sheet styled workbooks for EcoVerzz AI.
    Worksheets: Summary, Reports, Users, Categories, Leaderboard.
    """

    @classmethod
    def generate_excel(
        cls,
        output_filepath: str,
        analytics: dict,
        categories: list,
        top_users: list,
        monthly_data: list,
    ) -> str:
        wb = openpyxl.Workbook()
        # Default sheet -> Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_reports = wb.create_sheet(title="Reports")
        ws_users = wb.create_sheet(title="Users")
        ws_categories = wb.create_sheet(title="Categories")
        ws_leaderboard = wb.create_sheet(title="Leaderboard")

        # Styling Definitions
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        dark_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        accent_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="10B981")
        bold_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        regular_font = Font(name="Calibri", size=11, color="334155")

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # ---------------- 1. SUMMARY SHEET ---------------- #
        ws_summary.append(["EcoVerzz AI - Executive Summary Report"])
        ws_summary.cell(row=1, column=1).font = title_font
        ws_summary.append([])

        ws_summary.append(["Metric Parameter", "Current Value"])
        ws_summary.cell(row=3, column=1).fill = header_fill
        ws_summary.cell(row=3, column=1).font = header_font
        ws_summary.cell(row=3, column=2).fill = header_fill
        ws_summary.cell(row=3, column=2).font = header_font

        summary_rows = [
            ["Total Reports Submitted", analytics.get("total_reports", 0)],
            ["Pending Verification", analytics.get("pending", 0)],
            ["Verified Reports", analytics.get("verified", 0)],
            ["Resolved Reports", analytics.get("resolved", 0)],
            ["Total Eco Points Issued", analytics.get("eco_points", 0)],
        ]
        for r in summary_rows:
            ws_summary.append(r)

        # ---------------- 2. REPORTS SHEET ---------------- #
        ws_reports.append(["Month", "Total Reports", "Resolved Reports", "Eco Points"])
        for m in monthly_data:
            ws_reports.append([m["month"], m["total_reports"], m["resolved_reports"], m["eco_points"]])

        # ---------------- 3. USERS SHEET ---------------- #
        ws_users.append(["User ID", "Full Name", "Email Address", "Reports", "Eco Points"])
        for u in top_users:
            ws_users.append([u["user_id"], u["user_name"], u["email"], u["total_reports"], u["eco_points"]])

        # ---------------- 4. CATEGORIES SHEET ---------------- #
        ws_categories.append(["Category Name", "Report Count", "Eco Points Issued"])
        for c in categories:
            ws_categories.append([c["category"], c["count"], c["eco_points"]])

        # ---------------- 5. LEADERBOARD SHEET ---------------- #
        ws_leaderboard.append(["Rank", "Contributor Name", "Email", "Total Eco Points"])
        for idx, u in enumerate(top_users, 1):
            ws_leaderboard.append([idx, u["user_name"], u["email"], u["eco_points"]])

        # Apply Styling & Auto-Fit to all sheets
        for sheet in wb.worksheets:
            # Freeze top row
            sheet.freeze_panes = "A2"

            # Style header row (row 1 unless summary)
            start_r = 3 if sheet.title == "Summary" else 1
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=start_r, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply borders & font to data cells
            for row in sheet.iter_rows(min_row=start_r + 1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = regular_font
                    cell.border = thin_border

            # Auto-fit column widths
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_filepath)
        return output_filepath
